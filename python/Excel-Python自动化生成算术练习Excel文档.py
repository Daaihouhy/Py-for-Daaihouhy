import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font

def generate_expression():
    num1 = random.randint(1, 100)
    num2 = random.randint(1, 50)
    operator = '-' if num1 > num2 else '+'
    return f"{num1} {operator} {num2}="
# 以下为四则运算修改的代码
'''
def generate_expression():
    operators = ['+', '-', '×', '÷']
    operator = random.choice(operators)
    if operator in ['+', '-']:
        num1 = random.randint(1, 100)
        num2 = random.randint(1, 50)
    elif operator == '×':
        num1 = random.randint(1, 10)
        num2 = random.randint(1, 10)
    elif operator == '÷':
        num1 = random.randint(2, 100)
        num2 = random.randint(1, num1)
        while num1 % num2 != 0:
            num1 = random.randint(2, 100)
            num2 = random.randint(1, num1)
    return f"{num1} {operator} {num2}="
'''

expressions = [generate_expression() for _ in range(120)]

df = pd.DataFrame({
    'A': expressions[:30],
    'B': None,
    'C': None,
    'D': expressions[30:60],
    'E': None,
    'F': None,
    'G': expressions[60:90],
    'H': None,
    'I': None,
    'J': expressions[90:],
})

file_path = r'输出路径\算术练习题.xlsx'
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, header=False)

wb = load_workbook(file_path)
ws = wb.active

ws.insert_rows(1)

bold_font = Font(bold=True, size=16)
ws['A1'] = '算术练习'
ws['A1'].font = bold_font
ws['J1'] = '姓名：'
ws['J1'].font = bold_font

wb.save(file_path)

print("数据已成功导出到Excel文件。")
